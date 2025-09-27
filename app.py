import os
import re
import threading
import time
import json
import sqlite3
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple

from flask import Flask, jsonify, request, render_template, Response, stream_with_context, make_response, send_from_directory
import pandas as pd
import requests


app = Flask(__name__, static_folder="static", template_folder="templates")


# -----------------------------
# Config
# -----------------------------
BASE_DIR = os.path.abspath(os.getcwd())
TXT_EXTENSIONS = {".txt"}
EXCEL_EXTENSIONS = {".xlsx", ".xls"}
DATA_DIR = os.path.join(BASE_DIR, "data_sentence")
SQLITE_DB_PATH = os.path.join(DATA_DIR, "coca.sqlite")
STATE_FILE_PATH = os.path.join(DATA_DIR, "loading_state.json")

# -----------------------------
# Chat Config
# -----------------------------
DATA_CHAT_DIR = os.path.join(BASE_DIR, "data_chat")
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
CHAT_SQLITE_PATH = os.path.join(DATA_CHAT_DIR, "chat.sqlite")
COOKIE_NAME_PREFIX = "chat_"

os.makedirs(DATA_CHAT_DIR, exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Allowed file types for uploads
ALLOWED_EXTENSIONS = {
    "image": {"jpg", "jpeg", "png", "gif", "bmp", "webp"},
    "video": {"mp4", "avi", "mov", "wmv", "flv", "webm"},
    "audio": {"mp3", "wav", "flac", "aac", "ogg"},
    "document": {"pdf", "doc", "docx", "txt", "rtf"},
    "archive": {"zip", "rar", "7z", "tar", "gz"},
}
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB

# In-memory state for chat
online_users: Dict[str, Dict[str, Any]] = {}

# Locks and cache
file_locks = {
    "messages": threading.RLock(),
    "version": threading.RLock(),
}

message_cache: Dict[str, Any] = {
    "version": 0,
}

def _chat_init_db() -> None:
    con = sqlite3.connect(CHAT_SQLITE_PATH)
    try:
        cur = con.cursor()
        # Pragmas for durability + reasonable performance
        cur.execute("PRAGMA journal_mode=WAL;")
        cur.execute("PRAGMA synchronous=NORMAL;")
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS chat_messages (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              content TEXT NOT NULL
            );
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS chat_meta (
              key TEXT PRIMARY KEY,
              value TEXT
            );
            """
        )
        # Initialize version
        cur.execute("INSERT OR IGNORE INTO chat_meta(key, value) VALUES('version','0')")
        con.commit()
    finally:
        con.close()

_chat_init_db()

def _allowed_file(filename: str) -> bool:
    if "." not in filename:
        return False
    ext = filename.rsplit(".", 1)[1].lower()
    for exts in ALLOWED_EXTENSIONS.values():
        if ext in exts:
            return True
    return False

def _get_file_type(filename: str) -> str:
    if "." not in filename:
        return "other"
    ext = filename.rsplit(".", 1)[1].lower()
    for file_type, exts in ALLOWED_EXTENSIONS.items():
        if ext in exts:
            return file_type
    return "other"

def _unique_filename(filename: str) -> str:
    name, ext = os.path.splitext(filename)
    return f"{int(time.time()*1000)}_{os.getpid()}{ext}"

def _get_current_version() -> int:
    with file_locks["version"]:
        try:
            con = sqlite3.connect(CHAT_SQLITE_PATH)
            cur = con.cursor()
            cur.execute("SELECT value FROM chat_meta WHERE key='version'")
            row = cur.fetchone()
            con.close()
            version = int(row[0]) if row and row[0] is not None else 0
            message_cache["version"] = version
            return version
        except Exception:
            return 0

def _increment_version() -> int:
    with file_locks["version"]:
        cur_v = _get_current_version()
        new_v = cur_v + 1
        con = sqlite3.connect(CHAT_SQLITE_PATH)
        try:
            c = con.cursor()
            c.execute("UPDATE chat_meta SET value=? WHERE key='version'", (str(new_v),))
            con.commit()
        finally:
            con.close()
        message_cache["version"] = new_v
        return new_v

def _rand_nick() -> str:
    adjectives = [
        "雀跃", "沉思", "慵懒", "俏皮", "优雅", "狂野", "内敛", "天真", "狡黠",
        "矜持", "奔放", "恬淡", "热烈", "疏离", "缠绵", "激昂", "颓废", "通透", "迷离",
        "空灵", "炽热", "温润", "冷冽", "璀璨", "朦胧", "锐利", "柔和", "深邃",
    ]
    nouns = [
        "猫", "狗", "兔", "熊", "狼", "貘", "鸭嘴兽", "犰狳", "树懒", "蜜獾",
        "食蚁兽", "狐獴", "水豚", "指猴", "袋熊", "麒麟", "凤凰", "独角兽", "狮鹫", "龙猫",
        "儒艮", "海天使", "翻车鱼", "灯塔水母", "叶海龙", "极乐鸟", "犀鸟", "鲸头鹳", "几维鸟", "冠蕉鹃",
        "蓝宝石螳螂虾", "吉丁虫", "蓝闪蝶", "长颈鹿象鼻虫", "彩虹锹甲",
    ]
    import random
    return random.choice(adjectives) + random.choice(nouns)

def _get_token() -> str:
    import random
    return str(int(time.time()*1000)) + str(random.randint(1000, 9999))

def _add_system_message(msg: str) -> None:
    add_message({"type": "sys", "msg": f"<span class=\"tips-warning\">{msg}</span>"})

def _update_online_status() -> None:
    now = time.time()
    timeout_keys: List[str] = []
    for uk, info in list(online_users.items()):
        if now - info.get("last_active", 0) > 900:
            timeout_keys.append(uk)
    for uk in timeout_keys:
        username = online_users.get(uk, {}).get("name", "")
        online_users.pop(uk, None)
        _add_system_message(f"<strong>{username}</strong>已超时退出")
        _add_system_message(f"<span class=\"tips-warning\">当前在线人数：{len(online_users)}</span>")
        _increment_version()

def _clear_uploads() -> None:
    try:
        for filename in os.listdir(UPLOAD_DIR):
            fp = os.path.join(UPLOAD_DIR, filename)
            try:
                if os.path.isfile(fp):
                    os.remove(fp)
            except Exception:
                pass
    except FileNotFoundError:
        pass

def chat_total_messages() -> int:
    con = sqlite3.connect(CHAT_SQLITE_PATH)
    try:
        cur = con.cursor()
        cur.execute("SELECT COUNT(*) FROM chat_messages")
        (cnt,) = cur.fetchone()
        return int(cnt or 0)
    finally:
        con.close()

def chat_fetch_messages(offset: int, limit: int = 1000) -> List[str]:
    con = sqlite3.connect(CHAT_SQLITE_PATH)
    try:
        cur = con.cursor()
        cur.execute("SELECT content FROM chat_messages ORDER BY id ASC LIMIT ? OFFSET ?", (int(limit), int(offset)))
        rows = cur.fetchall()
        return [r[0] for r in rows]
    finally:
        con.close()

def add_message(message_obj: Dict[str, Any]) -> int:
    content = json.dumps(message_obj, ensure_ascii=False, separators=(",", ":"))
    con = sqlite3.connect(CHAT_SQLITE_PATH)
    try:
        cur = con.cursor()
        cur.execute("INSERT INTO chat_messages(content) VALUES(?)", (content,))
        con.commit()
        # cap to latest 10000
        cur.execute("SELECT COUNT(*) FROM chat_messages")
        (cnt,) = cur.fetchone()
        surplus = int(cnt or 0) - 10000
        if surplus > 0:
            cur.execute(
                "DELETE FROM chat_messages WHERE id IN (SELECT id FROM chat_messages ORDER BY id ASC LIMIT ?)",
                (surplus,),
            )
            con.commit()
        return int(cnt or 0)
    finally:
        con.close()

def clear_messages() -> None:
    con = sqlite3.connect(CHAT_SQLITE_PATH)
    try:
        cur = con.cursor()
        cur.execute("DELETE FROM chat_messages")
        con.commit()
    finally:
        con.close()


class LoadingStateStore:
    DEFAULT_STATE: Dict[str, Any] = {
        "running": False,
        "file": None,
        "current_sheet": None,
        "processed_words": 0,
        "total_words": 0,
        "percent": 0.0,
        "error": None,
        "latest_words": [],
        "timestamp": None,
    }

    def __init__(self, path: str):
        self.path = path
        self.lock = threading.Lock()
        self.state: Dict[str, Any] = self.DEFAULT_STATE.copy()
        self._last_persist = 0.0
        self._load_from_disk()

    def _load_from_disk(self) -> None:
        if not os.path.exists(self.path):
            # ensure directory exists for future writes
            os.makedirs(os.path.dirname(self.path), exist_ok=True)
            self._persist_locked(force=True)
            return
        try:
            with open(self.path, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            if isinstance(data, dict):
                merged = self.DEFAULT_STATE.copy()
                for key in merged.keys():
                    if key in data:
                        merged[key] = data[key]
                if not isinstance(merged.get("latest_words"), list):
                    merged["latest_words"] = []
                self.state = merged
        except Exception:
            # fall back to default when corrupted
            self.state = self.DEFAULT_STATE.copy()
        self._persist_locked(force=True)

    def _persist_locked(self, force: bool = False) -> None:
        now = time.time()
        if not force and (now - self._last_persist) < 0.5:
            return
        os.makedirs(os.path.dirname(self.path), exist_ok=True)
        tmp_path = f"{self.path}.tmp"
        with open(tmp_path, "w", encoding="utf-8") as fh:
            json.dump(self.state, fh, ensure_ascii=False)
        os.replace(tmp_path, self.path)
        self._last_persist = now

    def snapshot(self) -> Dict[str, Any]:
        with self.lock:
            return json.loads(json.dumps(self.state))

    def reset_for_file(self, file_name: str, total_rows: int) -> None:
        with self.lock:
            self.state = self.DEFAULT_STATE.copy()
            self.state.update({
                "running": True,
                "file": file_name,
                "current_sheet": None,
                "processed_words": 0,
                "total_words": int(total_rows or 0),
                "percent": 0.0,
                "error": None,
                "latest_words": [],
                "timestamp": datetime.utcnow().isoformat(),
            })
            self._persist_locked(force=True)

    def set_current_sheet(self, sheet: Optional[str]) -> None:
        with self.lock:
            self.state["current_sheet"] = sheet
            self.state["timestamp"] = datetime.utcnow().isoformat()
            self._persist_locked()

    def increment_processed(self, increment: int = 1, sample_word: Optional[str] = None,
                             sample_step: int = 10, latest_limit: int = 40, force: bool = False) -> int:
        with self.lock:
            processed = self.state.get("processed_words", 0) + increment
            self.state["processed_words"] = processed
            total = self.state.get("total_words", 0) or 0
            self.state["percent"] = (processed / total * 100.0) if total > 0 else 0.0
            if sample_word and sample_word.strip():
                if processed % sample_step == 0:
                    latest = list(self.state.get("latest_words") or [])
                    latest.append(sample_word)
                    if len(latest) > latest_limit:
                        latest = latest[-latest_limit:]
                    self.state["latest_words"] = latest
            self.state["timestamp"] = datetime.utcnow().isoformat()
            self._persist_locked(force=force)
            return processed

    def mark_finished(self, error: Optional[str] = None) -> None:
        with self.lock:
            self.state["running"] = False
            if error:
                self.state["error"] = error
            self.state["timestamp"] = datetime.utcnow().isoformat()
            self._persist_locked(force=True)

    def clear_error(self) -> None:
        with self.lock:
            self.state["error"] = None
            self.state["timestamp"] = datetime.utcnow().isoformat()
            self._persist_locked()


loading_state = LoadingStateStore(STATE_FILE_PATH)


# -----------------------------
# Current selected Excel file
# -----------------------------
current_excel_file: Optional[str] = None

# Loading state
loading_thread: Optional[threading.Thread] = None
loading_cancelled: bool = False


def normalize_word(word: str) -> str:
    if word is None:
        return ""
    # Keep letters, apostrophes, hyphens; lower-case
    text = str(word).strip()
    text = re.sub(r"[^A-Za-z\-']+", " ", text).strip().lower()
    return text


def list_excel_files() -> List[Dict[str, Any]]:
    files: List[Dict[str, Any]] = []
    for name in os.listdir(BASE_DIR):
        path = os.path.join(BASE_DIR, name)
        if os.path.isfile(path):
            _, ext = os.path.splitext(name)
            if ext.lower() in EXCEL_EXTENSIONS:
                stat = os.stat(path)
                files.append({
                    "name": name,
                    "size": stat.st_size,
                    "mtime": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                })
    files.sort(key=lambda x: x["name"].lower())
    return files

 


def compute_total_rows(file_path: str) -> int:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        total = 0
        for ws in wb.worksheets:
            total += (ws.max_row or 0)
        return total
    except Exception:
        # Fallback: estimate by parsing sheets with pandas (slower)
        try:
            xls = pd.ExcelFile(file_path)
            total = 0
            for sheet in xls.sheet_names:
                df = xls.parse(sheet_name=sheet, header=None, dtype=str)
                total += df.shape[0]
            return total
        except Exception:
            return 0


def _rebuild_sqlite_from_excel(file_path: str) -> None:
    """Rebuild a single SQLite database from the given Excel file.
    - During rebuild, querying is disabled (handled by caller via flags)
    - After successful rebuild, mark DATA_LOADED=True
    Table schema columns: word_norm, word, phonetic, meaning, sheet, row_index
    """
    # Prepare excel reader (streaming)
    from openpyxl import load_workbook
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    sheets = [ws.title for ws in wb.worksheets]

    # Create directory for DB if missing
    os.makedirs(DATA_DIR, exist_ok=True)

    # Create SQLite and write in one transaction
    con = sqlite3.connect(SQLITE_DB_PATH)
    cur = con.cursor()
    try:
        # Pragmas for faster build
        cur.execute("PRAGMA journal_mode=WAL;")
        cur.execute("PRAGMA synchronous=NORMAL;")
        cur.execute("PRAGMA temp_store=MEMORY;")

        cur.execute("DROP TABLE IF EXISTS entries;")
        cur.execute(
            """
            CREATE TABLE entries (
              id INTEGER PRIMARY KEY,
              word_norm TEXT NOT NULL,
              word TEXT,
              phonetic TEXT,
              meaning TEXT,
              sheet TEXT,
              row_index INTEGER
            );
            """
        )

        con.execute("BEGIN;")

        # Sampling config for preview
        SAMPLE_STEP = 10
        LATEST_LIMIT = 40
        loading_state.clear_error()

        insert_sql = (
            "INSERT INTO entries (word_norm, word, phonetic, meaning, sheet, row_index) VALUES (?, ?, ?, ?, ?, ?)"
        )

        # We reuse total_words computed by caller; processed_words is reset in caller
        for sheet in sheets:
            loading_state.set_current_sheet(sheet)
            ws = wb[sheet]
            # choose word column index based on worksheet column count
            max_cols = ws.max_column or 1
            word_col_idx = 1 if max_cols > 1 else 0
            # iterate and batch insert
            batch: List[Tuple[str, Optional[str], Optional[str], Optional[str], str, int]] = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if loading_cancelled:
                    con.rollback()
                    raise RuntimeError("loading cancelled")
                # guard for row tuple shorter than expected
                word_raw = row[word_col_idx] if word_col_idx < len(row or ()) else None
                display_word = "" if word_raw is None else str(word_raw).strip()
                norm = normalize_word(display_word)
                # map optional columns 2 and 3 when present
                phonetic_val = row[2] if (row and len(row) > 2) else None
                meaning_val = row[3] if (row and len(row) > 3) else None
                phonetic = None if phonetic_val is None else str(phonetic_val)
                meaning = None if meaning_val is None else str(meaning_val)

                batch.append((norm, display_word or None, phonetic, meaning, sheet, int(row_idx)))

                processed_now = loading_state.increment_processed(
                    increment=1,
                    sample_word=display_word,
                    sample_step=SAMPLE_STEP,
                    latest_limit=LATEST_LIMIT,
                )

                # flush in batches
                if len(batch) >= 10000:
                    cur.executemany(insert_sql, batch)
                    batch.clear()
                    # update percent after a chunk

            if batch:
                cur.executemany(insert_sql, batch)
                batch.clear()

        # index after all inserts
        cur.execute("CREATE INDEX idx_entries_word_norm ON entries(word_norm);")
        # speed up row lookup by (sheet,row_index)
        cur.execute("CREATE INDEX idx_entries_sheet_row ON entries(sheet, row_index);")

        con.commit()

        # mark loaded
        app.config["DATA_LOADED"] = True
        global current_excel_file
        current_excel_file = file_path
    finally:
        con.close()

 


def _loader_worker(file_path: str):
    global loading_thread, loading_cancelled
    try:
        file_name = os.path.basename(file_path)
        total_rows = compute_total_rows(file_path)
        loading_state.reset_for_file(file_name, total_rows)
        loading_cancelled = False
        _rebuild_sqlite_from_excel(file_path)
    except Exception as exc:
        loading_state.mark_finished(error=str(exc))
    finally:
        loading_state.mark_finished()
        loading_thread = None


def list_txt_files() -> List[str]:
    files: List[str] = []
    if not os.path.isdir(DATA_DIR):
        return files
    for name in os.listdir(DATA_DIR):
        path = os.path.join(DATA_DIR, name)
        if os.path.isfile(path):
            _, ext = os.path.splitext(name)
            if ext.lower() in TXT_EXTENSIONS:
                files.append(name)
    # Sort by natural order when possible
    files.sort(key=lambda s: [int(t) if t.isdigit() else t.lower() for t in re.split(r"(\d+)", s)])
    return files


# -----------------------------
# Routes
# -----------------------------
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/txt/list")
def api_txt_list():
    return jsonify({"files": list_txt_files()})


@app.route("/api/txt/content")
def api_txt_content():
    name = request.args.get("name", "").strip()
    if not name:
        return jsonify({"error": "missing name"}), 400
    # Security: only allow files from base dir and with .txt
    safe_name = os.path.basename(name)
    _, ext = os.path.splitext(safe_name)
    if ext.lower() not in TXT_EXTENSIONS:
        return jsonify({"error": "invalid file type"}), 400
    base = DATA_DIR if os.path.isdir(DATA_DIR) else BASE_DIR
    file_path = os.path.join(base, safe_name)
    if not os.path.exists(file_path):
        return jsonify({"error": "file not found"}), 404
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
        return jsonify({"name": safe_name, "content": content})
    except UnicodeDecodeError:
        with open(file_path, "r", encoding="gb18030", errors="ignore") as f:
            content = f.read()
        return jsonify({"name": safe_name, "content": content})


@app.route("/api/excel/files")
def api_excel_files():
    return jsonify({"files": list_excel_files()})


@app.route("/api/excel/status")
def api_excel_status():
    # 简化：仅以数据库是否存在且包含 entries 表为准
    actual_ready = False
    try:
        if os.path.exists(SQLITE_DB_PATH):
            con = sqlite3.connect(SQLITE_DB_PATH)
            cur = con.cursor()
            cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='entries'")
            actual_ready = cur.fetchone() is not None
            con.close()
    except Exception:
        actual_ready = False
    state = loading_state.snapshot()
    state.update({
        "loaded": bool(actual_ready),
        "current_file": current_excel_file,
    })
    return jsonify(state)


@app.route("/api/excel/stream")
def api_excel_stream():
    def generate():
        try:
            max_seconds = float(request.args.get("duration", 10.0))
        except ValueError:
            max_seconds = 10.0
        max_seconds = max(1.0, min(max_seconds, 60.0))

        try:
            interval = float(request.args.get("interval", 0.5))
        except ValueError:
            interval = 0.5
        interval = max(0.1, min(interval, 2.0))

        last_sent: Dict[str, Any] = {}
        last_emit_ts = 0.0
        start_time = time.time()

        while True:
            snapshot = loading_state.snapshot()
            state = {
                "loaded": bool(app.config.get("DATA_LOADED", False)),
                **snapshot,
            }

            changed = False
            if not last_sent:
                changed = True
            else:
                if state.get("timestamp") != last_sent.get("timestamp"):
                    changed = True
                elif int(state.get("percent", -1)) != int(last_sent.get("percent", -1)):
                    changed = True
                elif state.get("running") != last_sent.get("running"):
                    changed = True

            now = time.time()
            if changed or (now - last_emit_ts) >= max(2.0, interval * 3):
                last_sent = state
                last_emit_ts = now
                yield f"data: {json.dumps(state, ensure_ascii=False)}\n\n"

            if now - start_time >= max_seconds:
                break

            time.sleep(interval)

        closing_payload = {"event": "done", "timestamp": datetime.utcnow().isoformat()}
        yield f"data: {json.dumps(closing_payload, ensure_ascii=False)}\n\n"

    headers = {"Cache-Control": "no-cache", "X-Accel-Buffering": "no"}
    return Response(stream_with_context(generate()), mimetype="text/event-stream", headers=headers)


@app.route("/api/excel/load", methods=["POST"])
def api_excel_load():
    global loading_thread, current_excel_file
    file_name = request.args.get("file", "").strip()
    files = list_excel_files()
    allowed = {f["name"] for f in files}
    if not file_name or file_name not in allowed:
        return jsonify({"error": "invalid file"}), 400
    if loading_state.snapshot().get("running"):
        return jsonify({"error": "loading in progress"}), 409
    # reset state
    app.config["DATA_LOADED"] = False
    current_excel_file = None
    # remove existing sqlite db if any (fresh rebuild as requested)
    try:
        if os.path.exists(SQLITE_DB_PATH):
            os.remove(SQLITE_DB_PATH)
    except Exception:
        pass
    # clear legacy in-memory structures (no longer used)
    # start thread
    file_path = os.path.join(BASE_DIR, file_name)
    t = threading.Thread(target=_loader_worker, args=(file_path,), daemon=True)
    loading_thread = t
    t.start()
    return jsonify({"started": True})


@app.route("/api/excel/unload", methods=["POST"])
def api_excel_unload():
    global current_excel_file
    app.config["DATA_LOADED"] = False
    current_excel_file = None
    # clear legacy in-memory structures (no longer used)
    # delete sqlite db file as well
    try:
        if os.path.exists(SQLITE_DB_PATH):
            os.remove(SQLITE_DB_PATH)
    except Exception:
        pass
    return jsonify({"ok": True})


@app.route("/api/ai/chat", methods=["POST"])
def api_ai_chat():
    try:
        data = request.get_json(silent=True) or {}
        api_key = (data.get("api_key") or "").strip()
        content = (data.get("content") or "").strip()
        model = (data.get("model") or "Qwen/QwQ-32B").strip()
        system = (data.get("system") or "").strip()
        if not api_key:
            return jsonify({"error": "missing api_key"}), 400
        if not content:
            return jsonify({"error": "missing content"}), 400
        url = "https://api.siliconflow.cn/v1/chat/completions"
        messages = []
        if system:
            messages.append({"role": "system", "content": system})
        messages.append({"role": "user", "content": content})
        payload = {
            "model": model,
            "messages": messages,
        }
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        }
        r = requests.post(url, json=payload, headers=headers, timeout=60)
        r.raise_for_status()
        dj = r.json()
        msg = None
        try:
            msg = dj.get("choices", [{}])[0].get("message", {}).get("content", None)
        except Exception:
            msg = None
        cleaned = "" if msg is None else str(msg).lstrip("\r\n")
        return jsonify({
            "message": cleaned,
            "raw": dj,
        })
    except requests.HTTPError as http_err:
        return jsonify({"error": f"http {http_err.response.status_code}", "detail": http_err.response.text}), 502
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500

@app.route("/api/excel/search")
def api_excel_search():
    if not app.config.get("DATA_LOADED", False):
        return jsonify({"error": "loading or db not ready"}), 400
    word = request.args.get("word", "").strip()
    if not word:
        return jsonify({"error": "missing word"}), 400
    norm = normalize_word(word)
    try:
        con = sqlite3.connect(SQLITE_DB_PATH)
        cur = con.cursor()
        cur.execute(
            "SELECT sheet, row_index FROM entries WHERE word_norm = ? LIMIT 1",
            (norm,),
        )
        rows = cur.fetchall()
        matches: List[Dict[str, Any]] = []
        for s, r in rows:
            matches.append({"sheet": s, "row_index": int(r) if r is not None else 0})
        con.close()
        return jsonify({"word": word, "normalized": norm, "count": len(matches), "matches": matches})
    except Exception as exc:
        return jsonify({"error": f"db error: {exc}"}), 500


@app.route("/api/lookup")
def api_lookup():
    if not app.config.get("DATA_LOADED", False):
        return jsonify({"error": "loading or db not ready"}), 400
    word = request.args.get("word", "").strip()
    if not word:
        return jsonify({"error": "missing word"}), 400
    norm = normalize_word(word)
    try:
        con = sqlite3.connect(SQLITE_DB_PATH)
        cur = con.cursor()
        cur.execute(
            "SELECT word, phonetic, meaning FROM entries WHERE word_norm = ? LIMIT 1",
            (norm,),
        )
        result = cur.fetchone()
        con.close()
        if not result:
            return jsonify({"error": "not found"}), 404
        w, phonetic, meaning = result
        row_obj = {
            "1": w or "",
            "2": phonetic or "",
            "3": meaning or "",
        }
        return jsonify({"word": word, "row": row_obj})
    except Exception as exc:
        return jsonify({"error": f"db error: {exc}"}), 500

@app.route("/api/excel/row")
def api_excel_row():
    if not app.config.get("DATA_LOADED", False):
        return jsonify({"error": "loading or db not ready"}), 400
    sheet = request.args.get("sheet", "").strip()
    try:
        row_index = int(request.args.get("row_index", "-1"))
    except ValueError:
        row_index = -1
    if not sheet or row_index < 0:
        return jsonify({"error": "missing sheet or row_index"}), 400
    try:
        con = sqlite3.connect(SQLITE_DB_PATH)
        cur = con.cursor()
        cur.execute(
            "SELECT word, phonetic, meaning FROM entries WHERE sheet = ? AND row_index = ?",
            (sheet, row_index),
        )
        result = cur.fetchone()
        con.close()
        if not result:
            return jsonify({"error": "not found"}), 404
        word_text, phonetic, meaning = result
        row_obj = {
            "1": word_text or "",
            "2": phonetic or "",
            "3": meaning or "",
        }
        return jsonify({
            "sheet": sheet,
            "row_index": row_index,
            "row": row_obj,
        })
    except Exception as exc:
        return jsonify({"error": f"db error: {exc}"}), 500


# -----------------------------
# Chat Routes (integrated)
# -----------------------------

@app.route("/uploads/<path:filename>")
def chat_uploaded_file(filename: str):
    return send_from_directory(UPLOAD_DIR, filename)


@app.route("/login", methods=["POST"])
def chat_login():
    data = request.get_json(silent=True) or {}
    nickname = (data.get("n") or "").strip()
    if not nickname:
        nickname = _rand_nick()
    user_key = _get_token()
    _update_online_status()
    online_users[user_key] = {"name": nickname, "last_active": time.time()}
    _add_system_message(f"<strong>{nickname}</strong>已加入")
    _add_system_message(f"<span class=\"tips-warning\">当前在线人数：{len(online_users)}</span>")
    resp = make_response(jsonify({
        "name": nickname,
        "key": user_key,
        "version": _get_current_version(),
    }))
    max_age = 90 * 24 * 3600
    resp.set_cookie(COOKIE_NAME_PREFIX + "name", nickname, max_age=max_age)
    resp.set_cookie(COOKIE_NAME_PREFIX + "key", user_key, max_age=max_age)
    return resp


@app.route("/logout", methods=["POST"])
def chat_logout():
    username = request.cookies.get(COOKIE_NAME_PREFIX + "name", "")
    user_key = request.cookies.get(COOKIE_NAME_PREFIX + "key", "")
    if user_key in online_users:
        online_users.pop(user_key, None)
        _add_system_message(f"<strong>{username}</strong>已退出")
        _add_system_message(f"<span class=\"tips-warning\">当前在线人数：{len(online_users)}</span>")
    return jsonify({"result": "success", "version": _get_current_version()})


@app.route("/heartbeat", methods=["POST"])
def chat_heartbeat():
    user_key = request.cookies.get(COOKIE_NAME_PREFIX + "key", "")
    if user_key in online_users:
        online_users[user_key]["last_active"] = time.time()
    return jsonify({"result": "success"})


@app.route("/send", methods=["POST"])
def chat_send():
    message = (request.form.get("msg") or "").strip()
    if not message:
        return jsonify({"result": "error", "message": "消息不能为空"}), 400
    username = request.cookies.get(COOKIE_NAME_PREFIX + "name", "匿名")
    user_key = request.cookies.get(COOKIE_NAME_PREFIX + "key", "")
    if user_key in online_users:
        online_users[user_key]["last_active"] = time.time()
    if message == "/rm127.0.0.1":
        clear_messages()
        _clear_uploads()
        new_v = _increment_version()
        _add_system_message("💥 已清空所有聊天记录与上传文件！")
        return jsonify({"result": "success", "version": new_v, "admin_clear": True})
    msg_obj = {
        "type": "msg",
        "name": username,
        "key": user_key,
        "msg": message[:1000],
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M'),
    }
    add_message(msg_obj)
    return jsonify({"result": "success", "version": _get_current_version()})


@app.route("/upload", methods=["POST"])
def chat_upload():
    username = request.cookies.get(COOKIE_NAME_PREFIX + "name", "匿名")
    user_key = request.cookies.get(COOKIE_NAME_PREFIX + "key", "")
    if user_key in online_users:
        online_users[user_key]["last_active"] = time.time()
    if 'files[]' not in request.files:
        return jsonify({'result': 'error', 'message': '没有选择文件'}), 400
    files = request.files.getlist('files[]')
    if not files or all(f.filename == '' for f in files):
        return jsonify({'result': 'error', 'message': '没有选择文件'}), 400
    uploaded_files = []
    for file in files:
        if not file or file.filename == '':
            continue
        if not _allowed_file(file.filename):
            continue
        file.seek(0, os.SEEK_END)
        size = file.tell()
        file.seek(0)
        if size > MAX_FILE_SIZE:
            return jsonify({'result': 'error', 'message': f'文件 {file.filename} 超过大小限制 (50MB)'}), 400
        from werkzeug.utils import secure_filename
        original_filename = secure_filename(file.filename)
        unique_filename = _unique_filename(original_filename)
        save_path = os.path.join(UPLOAD_DIR, unique_filename)
        file.save(save_path)
        file_info = {
            'name': original_filename,
            'filename': unique_filename,
            'size': size,
            'type': _get_file_type(original_filename),
        }
        msg = {
            'type': 'file',
            'name': username,
            'key': user_key,
            'fileInfo': file_info,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M'),
        }
        add_message(msg)
        uploaded_files.append(file_info)
    new_v = _increment_version()
    return jsonify({'result': 'success', 'files': uploaded_files, 'version': new_v})


@app.route("/msg", methods=["GET"])
def chat_get_messages():
    last_index = int(request.args.get('k', 0))
    client_version = int(request.args.get('v', 0))
    server_version = _get_current_version()
    _update_online_status()
    if client_version != server_version:
        return jsonify({'reset': True, 'version': server_version})
    total = chat_total_messages()
    if last_index >= total:
        return jsonify({'count': total, 'list': [], 'version': server_version})
    start_index = max(0, total - 1000)
    start_index = max(start_index, last_index)
    limit = min(1000, total - start_index)
    raw_list = chat_fetch_messages(start_index, limit)
    users = set()
    for s in raw_list:
        try:
            obj = json.loads(s)
            if isinstance(obj, dict) and obj.get('type') == 'msg' and obj.get('name'):
                users.add(obj['name'])
        except Exception:
            pass
    return jsonify({'count': total, 'list': raw_list, 'version': server_version, 'users': list(users)})


# No auto-loading. Data is loaded via /api/excel/load
pass


if __name__ == "__main__":
    # Auto-detect existing SQLite database on startup
    try:
        db_ready = False
        if os.path.exists(SQLITE_DB_PATH):
            con = sqlite3.connect(SQLITE_DB_PATH)
            cur = con.cursor()
            cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='entries'")
            exists = cur.fetchone() is not None
            con.close()
            if exists:
                app.config["DATA_LOADED"] = True
                db_ready = True
                # Best-effort set current excel file for UI display
                data_xlsx_path = os.path.join(BASE_DIR, "data.xlsx")
                if os.path.exists(data_xlsx_path):
                    current_excel_file = data_xlsx_path
        # If DB not ready, try to auto-load once on startup
        if not db_ready:
            excel_candidates = [
                name for name in os.listdir(BASE_DIR)
                if os.path.isfile(os.path.join(BASE_DIR, name))
                and os.path.splitext(name)[1].lower() in EXCEL_EXTENSIONS
            ]
            excel_candidates.sort()
            auto_excel = excel_candidates[0] if excel_candidates else None
            if auto_excel:
                auto_excel_path = os.path.join(BASE_DIR, auto_excel)
                is_main_worker = (os.environ.get("WERKZEUG_RUN_MAIN") == "true") or not bool(os.environ.get("WERKZEUG_RUN_MAIN"))
                if os.path.exists(auto_excel_path) and is_main_worker and not loading_state.snapshot().get("running"):
                    app.config["DATA_LOADED"] = False
                    try:
                        if os.path.exists(SQLITE_DB_PATH):
                            os.remove(SQLITE_DB_PATH)
                    except Exception:
                        pass
                    t = threading.Thread(target=_loader_worker, args=(auto_excel_path,), daemon=True)
                    loading_thread = t
                    t.start()
    except Exception:
        pass
# 调试模式，开发时用
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
    # 生产模式，部署时用
    # from waitress import serve

    # # 根据服务器配置(2vCPU/2GB内存)优化并发处理能力
    # serve(
    #     app, 
    #     host="0.0.0.0", 
    #     port=5000,
    #     threads=10,             # 适合2核CPU的线程数
    #     connection_limit=500,   # 适合2GB内存的连接数
    #     channel_timeout=120     # 连接超时时间(秒)
    # )


